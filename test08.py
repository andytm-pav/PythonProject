import PySimpleGUI as sg
import openpyxl
from django.template.defaultfilters import length
from openpyxl import Workbook
import os

# Файл с данными
mapping_file = "mapping.xlsx"
sheet_name = "mapping"


# Функция для загрузки данных из файла
def load_mapping_data():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return {}

    wb = openpyxl.load_workbook(mapping_file, data_only=True)

    if sheet_name not in wb.sheetnames:
        sg.popup_error(f"Лист '{sheet_name}' не найден в файле!")
        return {}

    ws = wb[sheet_name]

    # for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
    #     print(row)

    # Загружаем данные из колонок A (ключ) и B (значение), начиная с A2 и B2
    mapping_data = {row[0]: row[1] for row in ws.iter_rows(min_row=2, max_col=2, values_only=True) if
                    row[0]}
    # print(mapping_data)
    return mapping_data


# Загружаем данные
mapping_data = load_mapping_data()

print(mapping_data)

if not mapping_data:
    sg.popup_error("Ошибка: нет данных для загрузки!")
    exit()

# Определяем макет окна
layout = [
             [sg.Text(str(value) + ":" + '_'*(50-length(str(value)))), sg.InputText(default_text="", key=f"-INPUT_{key}-")] for key, value in
             mapping_data.items()
         ] + [
             [sg.Button("Сохранить", key="-SAVE-"), sg.Button("Выход", key="-EXIT-")]
         ]

# Создаем окно
window = sg.Window("Моя программа", layout, size=(1000, 500))

while True:
    event, values = window.read()

    if event in (sg.WINDOW_CLOSED, "-EXIT-"):
        break
    elif event == "-SAVE-":
        # Собираем данные в словарь
        data_dict = {key: values[f"-INPUT_{key}-"] for key in mapping_data.keys()}

        # Файл для сохранения данных
        file_path = "test.xlsx"

        # Проверяем, существует ли файл
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)  # Открываем существующий файл
            ws = wb.active
        else:
            wb = Workbook()  # Создаем новый файл
            ws = wb.active
            ws.append(list(data_dict.keys()))  # Добавляем заголовки, если файла нет

        # Добавляем новую строку с данными
        ws.append(list(data_dict.values()))

        # Сохраняем Excel-файл
        wb.save(file_path)

        sg.popup("Данные сохранены!", title="Успех")

window.close()
