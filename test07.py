import PySimpleGUI as sg
import openpyxl
from openpyxl import Workbook
import os

# Файл с названиями полей
mapping_file = "mapping.xlsx"
sheet_name = "mapping"


# Функция для загрузки названий полей из файла
def load_field_names():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return []

    wb = openpyxl.load_workbook(mapping_file, data_only=True)

    if sheet_name not in wb.sheetnames:
        sg.popup_error(f"Лист '{sheet_name}' не найден в файле!")
        return []

    ws = wb[sheet_name]

    # Загружаем все значения из колонки A, начиная с A2
    field_names = [cell.value for cell in ws["A"][1:] if cell.value]  # Пропускаем A1 и убираем пустые ячейки
    return field_names


# Загружаем названия полей
field_names = load_field_names()

print(type(field_names))

if not field_names:
    sg.popup_error("Ошибка: нет данных для загрузки полей!")
    exit()

# Определяем макет окна
layout = [
             [sg.Text(name + ":"), sg.InputText(key=f"-INPUT{i}-")] for i, name in enumerate(field_names)
         ] + [
             [sg.Button("Сохранить", key="-SAVE-"), sg.Button("Выход", key="-EXIT-")]
         ]

# Создаем окно
window = sg.Window("Моя программа", layout, size=(400, 350))

while True:
    event, values = window.read()

    if event in (sg.WINDOW_CLOSED, "-EXIT-"):
        break
    elif event == "-SAVE-":
        # Собираем данные в словарь
        data_dict = {name: values[f"-INPUT{i}-"] for i, name in enumerate(field_names)}

        # Файл для сохранения данных
        file_path = "test.xlsx"

        # Проверяем, существует ли файл
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)  # Открываем существующий файл
            ws = wb.active
        else:
            wb = Workbook()  # Создаем новый файл
            ws = wb.active
            ws.append(field_names)  # Добавляем заголовки, если файла нет

        # Добавляем новую строку с данными
        ws.append(list(data_dict.values()))

        # Сохраняем Excel-файл
        wb.save(file_path)

        sg.popup("Данные сохранены!", title="Успех")

window.close()
