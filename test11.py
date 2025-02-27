import PySimpleGUI as sg
import openpyxl
import pandas as pd
import os
from pprint import pprint

# Файл с данными
mapping_file = "mapping.xlsx"
sheet_name = "mapping"
output_file_path = "tttest.xlsx"

# Функция для загрузки данных в словарь
def load_mapping_data():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return {}

    wb = openpyxl.load_workbook(mapping_file, data_only=True)

    if sheet_name not in wb.sheetnames:
        sg.popup_error(f"Лист '{sheet_name}' не найден в файле!")
        return {}

    ws = wb[sheet_name]

    # Загружаем данные из колонок A (ключ) и B (значение), начиная с A2 и B2
    mapping_data = {row[0]: row[1] for row in ws.iter_rows(min_row=2, max_col=2, values_only=True) if row[0] and row[1]}
    return mapping_data

# Функция для загрузки данных в DataFrame
def load_dataframe():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return pd.DataFrame()

    df = pd.read_excel(mapping_file, sheet_name=sheet_name, dtype=str)

    if df.shape[1] > 1:
        df.drop(df.columns[1], axis=1, inplace=True)

    df.set_index(df.columns[0], inplace=True)

    return df

# Загружаем данные
mapping_data = load_mapping_data()
df = load_dataframe()
# pprint(df.columns)

if df.empty or not mapping_data:
    sg.popup_error("Ошибка: нет данных для загрузки!")
    exit()

# Создаем макет с прокруткой
input_fields = [
    [sg.InputText(default_text="", key=f"-INPUT_{key}-"), sg.Text(f"{value}".ljust(60))]
    if not str(key).startswith("P") else [sg.Text(f"{value}".ljust(60)), sg.Text("", size=(60,1))]
    for key, value in mapping_data.items()
]

layout = [
    [sg.Button("Открыть файл", key="-OPEN-", pad=((50, 10), (20, 20))),
     sg.Text("", size=(10, 1)),
     sg.Button("Сохранить", key="-SAVE-"),
     sg.Text("", size=(30, 1)),
     sg.Button("Выход", key="-EXIT-")],
    [sg.Column(input_fields, size=(1000, 500), scrollable=True, vertical_scroll_only=True)]  # Прокручиваемая область
]

# Создаем окно
window = sg.Window("Моя программа", layout, size=(1000, 600), resizable=True)

while True:
    event, values = window.read()

    if event in (sg.WINDOW_CLOSED, "-EXIT-"):
        break
    elif event == "-OPEN-":
        output_file_path = None
        output_file_path = sg.popup_get_file("Выберите файл",
                                     file_types=(("Excel файлы", "*.xlsx"), ("Все файлы", "*.xlsx")))
        if output_file_path:
            sg.popup(f"Вы выбрали: {output_file_path}")
    elif event == "-SAVE-":
        # Собираем обновленные данные в словарь
        data_dict = {key: values[f"-INPUT_{key}-"] for key in mapping_data.keys() if f"-INPUT_{key}-" in values}

        # Заполняем result с учетом обновленного data_dict
        result = [
            {"sheet_name": column, "cell": df.loc[idx, column], "new_value": data_dict.get(idx, None)}
            for column in df.columns  # Пропускаем первый столбец
            for idx in df.index  # Проходим по строкам (идентификаторам)
        ]

        # print("Data Dictionary:", data_dict)
        # print("Processed Data:", result)
        # pprint(result)


        wb = openpyxl.load_workbook(output_file_path)

        for i in result:
            # print(i)
            if not isinstance(i['cell'], float) and i["new_value"] != '':
                cells = i["cell"].split(",")
                for c in cells:
                    print(i["sheet_name"], 'cells:'+c.strip(), 'value:'+i["new_value"])

                    sheet_name = i["sheet_name"]
                    cell_address = c.strip()
                    new_value = i["new_value"]

                    # Проверяем, существует ли указанный лист
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        try:
                            ws[cell_address] = new_value  # Обновляем значение ячейки
                        except AttributeError:
                            print(sheet_name, cell_address, new_value)
                            raise
                    else:
                        # print(f"Warning: Лист '{sheet_name}' не найден.")
                        sg.popup_error(f"Warning: Лист '{sheet_name}' не найден.")

        # Сохраняем изменения в файле
        wb.save(output_file_path)
        wb.close()


        sg.popup("Данные сохранены!", title="Успех")

window.close()
