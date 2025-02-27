import PySimpleGUI as sg
import openpyxl
from openpyxl import Workbook
import os

# Файл с данными
mapping_file = "mapping.xlsx"
sheet_name = "mapping"

# Функция для загрузки данных из файла
def load_mapping_data():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return {}

    wb = openpyxl.load_workbook(mapping_file, data_only=True)

    if sheet_name not in wb.sheetnames:
        sg.popup_error(f"Лист '{sheet_name}' не найден в файле!")
        return {}

    ws = wb[sheet_name]

    # Загружаем данные из колонок A (ключ) и B (значение), начиная с A2 и B2
    mapping_data = {row[0]: row[1] for row in ws.iter_rows(min_row=2, max_col=2, values_only=True) if row[0] and row[1]}
    return mapping_data


# Загружаем данные
mapping_data = load_mapping_data()

if not mapping_data:
    sg.popup_error("Ошибка: нет данных для загрузки!")
    exit()

# Создаем макет с прокруткой
input_fields = [
    [sg.Text(f"{value}".ljust(60)), sg.InputText(default_text="", key=f"-INPUT_{key}-")]
    if not str(key).startswith("P") else [sg.Text(f"{value}".ljust(60)), sg.Text("", size=(60,1))]
    for key, value in mapping_data.items()
]

layout = [
    [sg.Column(input_fields, size=(800, 500), scrollable=True, vertical_scroll_only=True)],  # Прокручиваемая область
    [sg.Button("Сохранить", key="-SAVE-"), sg.Button("Выход", key="-EXIT-")]
]

# Создаем окно
window = sg.Window("Моя программа", layout, size=(1000, 600), resizable=True)

while True:
    event, values = window.read()

    if event in (sg.WINDOW_CLOSED, "-EXIT-"):
        break
    elif event == "-SAVE-":
        # Собираем данные в словарь (не включая скрытые поля)
        data_dict = {key: values[f"-INPUT_{key}-"] for key in mapping_data.keys() if f"-INPUT_{key}-" in values}

        print(data_dict)

        sg.popup("Данные сохранены!", title="Успех")

window.close()
