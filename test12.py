import PySimpleGUI as sg
import openpyxl
import pandas as pd
import os
import json
import shutil
import datetime
import re
from pprint import pprint


# Файл с данными
mapping_file = "mapping.xlsx"
sheet_name = "mapping"
src_xlsx_path = "tttest.xlsx"
data_json_path = "1.json"

# Функция для загрузки данных в словарь
def load_mapping_data():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return {}

    xwb = openpyxl.load_workbook(mapping_file, data_only=True)

    if sheet_name not in xwb.sheetnames:
        sg.popup_error(f"Лист '{sheet_name}' не найден в файле!")
        return {}

    xws = xwb[sheet_name]

    # Загружаем данные из колонок A (ключ) и B (значение), начиная с A2 и B2
    map_data = {row[0]: row[1] for row in xws.iter_rows(min_row=2, max_col=2, values_only=True) if row[0] and row[1]}
    return map_data

# Функция для загрузки данных в DataFrame
def load_dataframe():
    if not os.path.exists(mapping_file):
        sg.popup_error(f"Файл {mapping_file} не найден!")
        return pd.DataFrame()

    idf = pd.read_excel(mapping_file, sheet_name=sheet_name, dtype=str)

    if idf.shape[1] > 1:
        idf.drop(idf.columns[1], axis=1, inplace=True)

    idf.set_index(idf.columns[0], inplace=True)

    return idf

# Загружаем данные
mapping_data = load_mapping_data()
df = load_dataframe()
# pprint(df.columns)

if df.empty or not mapping_data:
    sg.popup_error("Ошибка: нет данных для загрузки!")
    exit()

# Создаем макет с прокруткой
input_fields = [
    [sg.InputText(default_text="", key=f"-INPUT_{key}-"), sg.Text(f"{value}".ljust(60))]
    if not str(key).startswith("P") else [sg.Text(f"{value}".ljust(60)), sg.Text("", size=(60,1))]
    for key, value in mapping_data.items()
]

layout = [
    [sg.Button("Загрузить данные для ввода JSON", key="-LOAD_JSON-", pad=((10, 10), (10, 10))),
     sg.Text("", size=(60,1), key="-JSON_PATH-"),
     sg.InputText(visible=False, enable_events=True, key='-SAVE_JSON_PATH-'),
     sg.FileSaveAs("Сохранить данные для ввода JSON", key='-SAVE_JSON-', file_types=(('JSON', '.json'),))
     ],
    [sg.Button("          Выбрать шаблон XLSX         ", key="-OPEN-", pad=((10, 10), (10, 10))),
     sg.Text("", size=(60,1), key="-XLSX_PATH-"),
     sg.Button("    Записать данные в файл XLSX     ", key="-SAVE-"),
     # sg.Text("", size=(30, 1)),
     # sg.Button("Выход", key="-EXIT-")
    ],
    [sg.Text("")],
    [sg.Column(input_fields, size=(1000, 500), scrollable=True, vertical_scroll_only=True)]  # Прокручиваемая область
]

# Создаем окно
window = sg.Window("Моя программа", layout, size=(1000, 600), resizable=True)

while True:
    event, values = window.read()

    if event in (sg.WINDOW_CLOSED, "-EXIT-"):
        break
    elif event == "-OPEN-":
        src_xlsx_path = sg.popup_get_file("Выберите файл XLSX", file_types=(("Excel файлы", "*.xlsx"),))
        if src_xlsx_path:
            # sg.popup(f"Вы выбрали: {src_xlsx_path}")
            window.Element("-XLSX_PATH-").Update(src_xlsx_path)

    elif event == "-LOAD_JSON-":
        data_json_path = sg.popup_get_file("Выберите файл JSON", file_types=(("JSON файлы", "*.json"),))
        if data_json_path:
            # sg.popup(f"Вы выбрали: {output_json_path}")
            window.Element("-JSON_PATH-").Update(data_json_path)
            with open(data_json_path, "r", encoding='utf-8') as file:
                data_dict = json.load(file)
                for key, value in data_dict.items():
                    window.Element(f"-INPUT_{key}-").Update(value)

    elif event == "-SAVE_JSON_PATH-":
        data_json_path = values['-SAVE_JSON_PATH-']
        window.Element("-JSON_PATH-").Update(data_json_path)
        data_dict = {key: values[f"-INPUT_{key}-"] for key in mapping_data.keys() if f"-INPUT_{key}-" in values}
        data_json = json.dumps(data_dict)
        with open(data_json_path, "w", encoding='utf-8') as file:  #
            file.write(json.dumps(json.loads(data_json), indent=4, sort_keys=False, ensure_ascii=False))  #
            sg.popup(f"Вы сохранили: {data_json_path}")

    elif event == "-SAVE-":
        # Собираем обновленные данные в словарь
        data_dict = {key: values[f"-INPUT_{key}-"] for key in mapping_data.keys() if f"-INPUT_{key}-" in values}

        # Заполняем result с учетом обновленного data_dict
        result = [
            {"sheet_name": column, "cell": df.loc[idx, column], "new_value": data_dict.get(idx, None)}
            for column in df.columns  # Пропускаем первый столбец
            for idx in df.index  # Проходим по строкам (идентификаторам)
        ]

        # print("Data Dictionary:", data_dict)
        # print("Processed Data:", result)
        # pprint(result)

        now = str(datetime.datetime.now())[:19]
        now = re.sub(r'[-: ]', "", now)

        dst_xlsx_path = re.sub(r'\.xlsx', "_" + str(now) + ".xlsx", src_xlsx_path)
        # print(src_xlsx_path, dst_xlsx_path)
        shutil.copy(src_xlsx_path, dst_xlsx_path)
        window.Element("-XLSX_PATH-").Update(dst_xlsx_path)

        wb = openpyxl.load_workbook(dst_xlsx_path)

        for i in result:
            # print(i)
            if not isinstance(i['cell'], float) and i["new_value"] != '':
                cells = i["cell"].strip(" ,").split(",")
                for c in cells:
                    # print(i["sheet_name"], 'cells:'+c.strip(), 'value:'+i["new_value"])

                    sheet_name = i["sheet_name"]
                    cell_address = c.strip()
                    new_value = i["new_value"]

                    # Проверяем, существует ли указанный лист
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        try:
                            ws[cell_address] = new_value  # Обновляем значение ячейки
                        except (AttributeError, IndexError):
                            print(sheet_name, cell_address, new_value)
                            sg.popup(f"Ошибка в маппинге: {sheet_name} - {cell_address} - {new_value}")
                            raise
                    else:
                        # print(f"Warning: Лист '{sheet_name}' не найден.")
                        sg.popup_error(f"Warning: Лист '{sheet_name}' не найден.")

        # Сохраняем изменения в файле
        wb.save(dst_xlsx_path)
        wb.close()

        sg.popup(f"Данные сохранены в '{dst_xlsx_path}' ", title="Успех")

window.close()
